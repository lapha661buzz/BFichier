import streamlit as st
from openpyxl import load_workbook
from io import BytesIO

TAILLE_MAX = 10 * 1024 # 10 Ko

st.set_page_config(page_title="Vérificateur Excel", page_icon="📊")

st.title("📂 Vérification de plusieurs fichiers Excel")
st.write("Chargez plusieurs fichiers `.xlsx`. Le total doit être inférieur à 10 Ko.")

# Uploader plusieurs fichiers
fichiers = st.file_uploader("📁 Déposez vos fichiers Excel ici", type=["xlsx"], accept_multiple_files=True)

if fichiers:
    taille_totale = sum(len(f.getvalue()) for f in fichiers)
    st.write(f"📦 Taille totale : `{taille_totale}` octets")
    st.write(f"📁 Nombre de fichiers : `{len(fichiers)}`")

    if taille_totale > TAILLE_MAX:
        st.error("❌ La taille totale dépasse 10 Ko.")
    else:
        st.success("✅ Tous les fichiers ont été acceptés (moins de 10 Ko au total).")
        for fichier in fichiers:
            try:
                wb = load_workbook(filename=BytesIO(fichier.read()), read_only=True)
                ws = wb.active
                nb_lignes = ws.max_row
                st.write(f"📄 `{fichier.name}` → `{nb_lignes}` lignes")
            except Exception as e:
                st.error(f"❌ Erreur dans `{fichier.name}` : {e}")